try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except ImportError:
    HAS_DND = False

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pandas as pd
import os
import sys
import re
import json
import webbrowser
import glob
import time
import threading
import urllib.request
from collections import defaultdict

GSHEET_URL = 'https://docs.google.com/spreadsheets/d/1boZBRpjPzvtjh2LM1K7f9aQT6kz5fLMBfEiSJ_rZaEc/export?format=xlsx&gid=661112978'
DOWNLOAD_FOLDER = os.path.expanduser('~/Downloads')

HAPPO_COLOR = 'FF66CCFF'

# 업데이트 관련 설정
GITHUB_REPO = 'wonderkjh10-cell/auto_run'
UPDATE_CHECK_URL = f'https://api.github.com/repos/{GITHUB_REPO}/releases/latest'

def _resource_path(relative_path):
    """PyInstaller 실행 시에도 동작하는 리소스 경로 반환"""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def get_current_version():
    """현재 프로그램 버전 읽기"""
    try:
        with open(_resource_path('version.txt'), 'r', encoding='utf-8') as f:
            return f.read().strip()
    except Exception:
        return '0.0.0'

def check_for_update():
    """GitHub에서 최신 버전 확인. (최신버전, 다운로드URL) 또는 None 반환."""
    try:
        req = urllib.request.Request(
            UPDATE_CHECK_URL,
            headers={'Accept': 'application/vnd.github.v3+json'}
        )
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))
        latest_version = data.get('tag_name', '').lstrip('v')
        current = get_current_version()
        if latest_version and latest_version != current:
            # 다운로드 URL 찾기
            download_url = None
            for asset in data.get('assets', []):
                if asset['name'].endswith('.exe'):
                    download_url = asset['browser_download_url']
                    break
            return (latest_version, download_url)
    except Exception:
        pass
    return None

MAPPING_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'mapping.json')

DEFAULT_MAPPING = {
    'ncp_1owiht_01': '아라올',
    'ehddk8080': '국민마트',
    'bh7555': '비에이치',
    'jjmall4817': '제이제이',
    'jjmall5031': '제이제이',
    'ltdcircle': '템스윈',
    'circlecir': '템스윈',
    'ltd.circle': '템스윈',
    'ltd_circle@naver.com': '템스윈',
    'circle1': '템스윈',
}


def load_saved_mapping():
    """저장된 매핑 파일 불러오기. 없으면 DEFAULT_MAPPING 반환"""
    if os.path.exists(MAPPING_FILE):
        try:
            with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return dict(DEFAULT_MAPPING)


def save_mapping_file(mapping):
    """매핑을 JSON 파일로 저장"""
    with open(MAPPING_FILE, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)


def is_happo(cell):
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
            return cell.fill.fgColor.rgb == HAPPO_COLOR
    except Exception:
        pass
    return False


def is_overseas(values, headers):
    """배송메세지에 '해외묶음번호'가 포함되면 해외배송으로 판별"""
    try:
        msg_col = headers.index('배송메세지')
        msg = str(values[msg_col] or '')
        return bool(re.search(r'해외묶음번호', msg))
    except (ValueError, IndexError):
        return False


def is_damaged(values, headers):
    """수집옵션명에 '훼손'이 포함되면 훼손건으로 판별"""
    try:
        opt_col = headers.index('수집옵션명')
        opt = str(values[opt_col] or '')
        return '훼손' in opt
    except (ValueError, IndexError):
        return False


def _parse_happo_rows(path):
    """xlsx를 직접 XML 파싱하여 합포(FF66CCFF) 행 번호 set 반환"""
    import zipfile
    import xml.etree.ElementTree as ET
    ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    happo_rows = set()
    try:
        z = zipfile.ZipFile(path)
        # fills에서 합포색 인덱스 찾기
        styles_xml = z.read('xl/styles.xml')
        root = ET.fromstring(styles_xml)
        fills = []
        for fill in root.findall(f'.//{{{ns}}}fills/{{{ns}}}fill'):
            pf = fill.find(f'{{{ns}}}patternFill')
            if pf is not None:
                fg = pf.find(f'{{{ns}}}fgColor')
                fills.append(fg.get('rgb') if fg is not None else None)
            else:
                fills.append(None)
        happo_fill_ids = {i for i, f in enumerate(fills) if f and f.upper() == HAPPO_COLOR}
        # cellXfs에서 합포 스타일 인덱스 찾기
        happo_styles = set()
        for i, xf in enumerate(root.findall(f'.//{{{ns}}}cellXfs/{{{ns}}}xf')):
            if int(xf.get('fillId', 0)) in happo_fill_ids:
                happo_styles.add(i)
        # sheet1.xml에서 각 행의 첫 셀 스타일 확인
        sheet_xml = z.read('xl/worksheets/sheet1.xml')
        sheet_root = ET.fromstring(sheet_xml)
        for row_el in sheet_root.findall(f'.//{{{ns}}}sheetData/{{{ns}}}row'):
            row_num = int(row_el.get('r'))
            cells = row_el.findall(f'{{{ns}}}c')
            if cells:
                style_idx = int(cells[0].get('s', '0'))
                if style_idx in happo_styles:
                    happo_rows.add(row_num)
        z.close()
    except Exception:
        pass
    return happo_rows


REQUIRED_ORDER_HEADERS = ['아이디', '수량', '상품코드', '사방넷 상품명', '배송메세지', '수집옵션명']


class OrderFileError(Exception):
    """발주서 파일 오류 (헤더 누락 등)"""
    pass


def _try_find_header_row(ws, max_scan=10):
    """시트에서 헤더 행 찾기. 탭 구분 데이터 자동 분리 지원.
    반환: (header_row_idx, clean_headers, split_by_tab) 또는 None"""
    rows_preview = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan), start=1):
        vals = [cell.value for cell in row]
        rows_preview.append((i, vals))

    for row_idx, vals in rows_preview:
        # 케이스 1: 일반 다열 헤더
        clean = [str(v).strip() if v is not None else '' for v in vals]
        if all(h in clean for h in REQUIRED_ORDER_HEADERS):
            return (row_idx, clean, False)

        # 케이스 2: 탭 구분 데이터가 A열에 몰린 경우
        if len(vals) >= 1 and vals[0] and isinstance(vals[0], str) and '\t' in vals[0]:
            split = [s.strip() for s in vals[0].split('\t')]
            if all(h in split for h in REQUIRED_ORDER_HEADERS):
                return (row_idx, split, True)

    return None


def load_order_file(path):
    """엄격 모드: 첫 시트의 1행을 헤더로만 인식. 깨진 파일이면 오류."""
    # XML에서 합포 행 번호 미리 파싱
    happo_rows = _parse_happo_rows(path)
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(), start=1):
        if i == 1:
            headers = [cell.value for cell in row]
            if headers is None:
                headers = []
            clean_headers = [str(h).strip() if h is not None else '' for h in headers]
            missing = [h for h in REQUIRED_ORDER_HEADERS if h not in clean_headers]
            if missing:
                wb.close()
                # 헤더 위치 표시: A열, B열 등
                def col_letter(idx):
                    return openpyxl.utils.get_column_letter(idx + 1)
                header_details = []
                for col_idx, h in enumerate(clean_headers):
                    if h:
                        header_details.append(f"  · {col_letter(col_idx)}열({col_idx + 1}열), 1행: '{h}'")
                # 시트 정보
                all_sheets = wb.sheetnames
                raise OrderFileError(
                    f"발주서 파일이 손상되었거나 올바른 양식이 아닙니다.\n\n"
                    f"[파일 정보]\n"
                    f"  · 시트 개수: {len(all_sheets)}개 {all_sheets}\n"
                    f"  · 현재 시트: '{ws.title}'\n"
                    f"  · 헤더(1행) 열 개수: {len([h for h in clean_headers if h])}개\n\n"
                    f"[현재 헤더 위치]\n"
                    + ('\n'.join(header_details) if header_details else '  · (없음)')
                    + f"\n\n"
                    f"[누락된 필수 항목]\n"
                    + '\n'.join(f"  · {h}" for h in missing)
                    + f"\n\n"
                    f"[해결방법]\n"
                    f"1. '자동복구' 버튼을 눌러 복구를 시도해보세요.\n"
                    f"2. 사방넷에서 발주서 파일을 다시 다운로드 받으세요.\n"
                    f"3. 다운로드 후 한셀 등 다른 프로그램으로 열지 말고 바로 사용하세요."
                )
            headers = clean_headers
            continue
        values = [cell.value for cell in row]
        rows.append({
            'values': values,
            'happo': i in happo_rows,
            'overseas': is_overseas(values, headers),
            'damaged': is_damaged(values, headers),
        })
    wb.close()
    return headers, rows


def recover_order_file(path):
    """깨진 발주서 파일에서 데이터 복구 시도.
    성공 시 복구된 파일 경로 반환. 실패 시 OrderFileError 발생.
    복구 파일은 원본과 같은 폴더에 '_복구됨' 접미사로 저장."""
    wb = openpyxl.load_workbook(path, read_only=True)

    target_ws = None
    header_row_idx = 1
    headers = None
    split_by_tab = False
    failed_sheets = []
    recovery_method = ''

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result = _try_find_header_row(ws)
        if result:
            header_row_idx, headers, split_by_tab = result
            target_ws = ws
            if sheet_name != wb.sheetnames[0]:
                recovery_method = f"다른 시트 '{sheet_name}'에서 데이터 발견"
            elif header_row_idx > 1:
                recovery_method = f"{header_row_idx}행에서 헤더 발견 (상단 {header_row_idx - 1}행 건너뜀)"
            elif split_by_tab:
                recovery_method = "A열의 탭 구분 데이터 분리"
            else:
                recovery_method = "정상 파일 (복구 불필요)"
            break
        else:
            try:
                first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
                preview = [cell.value for cell in first_row] if first_row else []
            except Exception:
                preview = []
            failed_sheets.append((sheet_name, preview))

    if target_ws is None:
        wb.close()
        preview_text = '\n'.join(
            f"  · 시트 '{name}' 첫 행: {p[:3] if len(p) <= 3 else p[:3] + ['...']}"
            for name, p in failed_sheets
        )
        raise OrderFileError(
            f"복구 실패: 발주서 형식을 찾을 수 없습니다.\n\n"
            f"[검사한 시트 ({len(failed_sheets)}개)]\n{preview_text}\n\n"
            f"필수 항목: {REQUIRED_ORDER_HEADERS}\n\n"
            f"사방넷에서 다시 다운로드 받으시는 것을 권장합니다."
        )

    # 복구 데이터 수집
    recovered_rows = []
    for row in target_ws.iter_rows(min_row=header_row_idx + 1):
        if split_by_tab:
            first_val = row[0].value if row and row[0].value else ''
            if not first_val or not isinstance(first_val, str):
                continue
            values = [s.strip() if s.strip() else None for s in first_val.split('\t')]
            values = [None if v == '' else v for v in values]
        else:
            values = [cell.value for cell in row]
            # 전부 None인 빈 행 스킵
            if all(v is None for v in values):
                continue
        recovered_rows.append(values)

    wb.close()

    # 복구된 데이터로 새 xlsx 저장
    base_dir = os.path.dirname(path)
    base_name = os.path.splitext(os.path.basename(path))[0]
    out_path = os.path.join(base_dir, f"{base_name}_복구됨.xlsx")

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = '복구됨'
    new_ws.append(headers)
    # 헤더 굵게
    for cell in new_ws[1]:
        cell.font = Font(bold=True)
    for values in recovered_rows:
        # 길이 맞추기
        if len(values) < len(headers):
            values = list(values) + [None] * (len(headers) - len(values))
        elif len(values) > len(headers):
            values = list(values)[:len(headers)]
        new_ws.append(values)
    new_wb.save(out_path)

    return out_path, recovery_method, len(recovered_rows)


def load_location_file(path):
    """위치 파일에서 상품코드 → 위치 매핑 반환"""
    df = pd.read_excel(path, sheet_name=0, header=0)
    location_map = {}
    for _, row in df.iterrows():
        code = str(row['상품코드']).strip() if pd.notna(row['상품코드']) else None
        loc  = str(row['위치']).strip()      if pd.notna(row['위치'])      else None
        if code and loc and code != 'nan' and loc != 'nan':
            location_map[code] = loc
    return location_map


def load_stock_file(path):
    df = pd.read_excel(path, sheet_name=0, header=None, skiprows=2)
    stock = {}
    for _, row in df.iterrows():
        code = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else None
        avail = row.iloc[16] if pd.notna(row.iloc[16]) else 0
        if code and code not in ('nan', '상품코드'):
            try:
                stock[code] = int(avail)
            except (ValueError, TypeError):
                stock[code] = 0
    return stock


def process_data(headers, rows, mapping, stock, location_map=None):
    id_col = headers.index('아이디')
    qty_col = headers.index('수량')
    code_col = headers.index('상품코드')
    name_col = headers.index('사방넷 상품명')

    # ① 위치코드 먼저 사방넷 상품명 앞에 추가 (정렬 전)
    if location_map:
        for row in rows:
            code = str(row['values'][code_col]).strip() if row['values'][code_col] else None
            if code and code in location_map:
                original = str(row['values'][name_col] or '')
                row['values'][name_col] = f"{location_map[code]} {original}"

    # 전체 상품코드별 총 수량 (모든 시트 합산, 유형별 집계)
    total_qty = defaultdict(int)
    total_normal_qty = defaultdict(int)
    total_happo_qty = defaultdict(int)
    total_overseas_qty = defaultdict(int)
    total_damaged_qty = defaultdict(int)
    for row in rows:
        code = str(row['values'][code_col]).strip() if row['values'][code_col] else None
        qty = int(row['values'][qty_col] or 0)
        if code:
            total_qty[code] += qty
            if row['overseas']:
                total_overseas_qty[code] += qty
            elif row['damaged']:
                total_damaged_qty[code] += qty
            elif row['happo']:
                total_happo_qty[code] += qty
            else:
                total_normal_qty[code] += qty

    # 아이디별 시트 분리
    sheets = defaultdict(list)
    for row in rows:
        row_id = str(row['values'][id_col]).strip() if row['values'][id_col] else '기타'
        sheet_name = mapping.get(row_id, row_id)
        sheets[sheet_name].append(row)

    result_sheets = {}
    for sheet_name, sheet_rows in sheets.items():
        # 정렬: 해외배송 → 훼손 → 합포 → 일반 (각 그룹 내 사방넷 상품명 오름차순)
        def sort_key(r):
            if r['overseas']:
                return 0
            if r['damaged']:
                return 1
            if r['happo']:
                return 2
            return 3
        sheet_rows = sorted(
            sheet_rows,
            key=lambda r: (
                sort_key(r),
                str(r['values'][name_col] or '').strip()
            )
        )

        normal_qty = defaultdict(int)
        happo_qty = defaultdict(int)
        overseas_qty = defaultdict(int)
        damaged_qty = defaultdict(int)

        for row in sheet_rows:
            code = str(row['values'][code_col]).strip() if row['values'][code_col] else None
            qty = int(row['values'][qty_col] or 0)
            if code:
                if row['overseas']:
                    overseas_qty[code] += qty
                elif row['damaged']:
                    damaged_qty[code] += qty
                elif row['happo']:
                    happo_qty[code] += qty
                else:
                    normal_qty[code] += qty

        seen_overseas = set()
        seen_damaged = set()
        seen_normal = set()
        seen_happo = set()
        result_rows = []
        for row in sheet_rows:
            code = str(row['values'][code_col]).strip() if row['values'][code_col] else None
            new_values = list(row['values'])

            if code and row['overseas'] and code not in seen_overseas:
                # 해외배송: 별개 제품 취급, 해외수량  0  잔여재고
                o = overseas_qty.get(code, 0)
                avail = stock.get(code, 0)
                # 잔여재고: 일반(합포 있으면 합포 제외) + 해외 + 훼손 차감
                tn = total_normal_qty.get(code, 0)
                th = total_happo_qty.get(code, 0)
                remaining = avail - (tn + th + total_overseas_qty.get(code, 0) + total_damaged_qty.get(code, 0))
                original = row['values'][name_col] or ''
                new_values[name_col] = f"{original}\n★★★      [해외] {o}   0   {remaining}      ★★★"
                seen_overseas.add(code)
            elif code and row['damaged'] and code not in seen_damaged:
                # 훼손: 별개 제품 취급, 훼손수량  0  잔여재고
                d = damaged_qty.get(code, 0)
                avail = stock.get(code, 0)
                tn = total_normal_qty.get(code, 0)
                th = total_happo_qty.get(code, 0)
                remaining = avail - (tn + th + total_overseas_qty.get(code, 0) + total_damaged_qty.get(code, 0))
                original = row['values'][name_col] or ''
                new_values[name_col] = f"{original}\n★★★      [훼손] {d}   0   {remaining}      ★★★"
                seen_damaged.add(code)
            elif code and not row['overseas'] and not row['damaged']:
                n = normal_qty.get(code, 0)
                h = happo_qty.get(code, 0)
                avail = stock.get(code, 0)
                tn = total_normal_qty.get(code, 0)
                th = total_happo_qty.get(code, 0)
                remaining = avail - (tn + th + total_overseas_qty.get(code, 0) + total_damaged_qty.get(code, 0))
                original = row['values'][name_col] or ''

                if row['happo'] and code not in seen_happo:
                    seen_happo.add(code)
                    if n == 0:
                        # 합포만 있는 경우: 0  합포수량  잔여재고 (합포 행에 표시)
                        new_values[name_col] = f"{original}\n★★★      0   {h}   {remaining}      ★★★"
                        seen_normal.add(code)
                    # 일반도 있는 경우: 합포 행에는 표기 안 함 (일반 행에서 표시)
                elif not row['happo'] and code not in seen_normal:
                    seen_normal.add(code)
                    if h == 0:
                        # 일반만 있는 경우: 일반수량  0  잔여재고
                        new_values[name_col] = f"{original}\n★★★      {n}   0   {remaining}      ★★★"
                    else:
                        # 일반+합포 있는 경우: (일반+합포)  합포수량  잔여재고 (일반 행에만 표시)
                        new_values[name_col] = f"{original}\n★★★      {n + h}   {h}   {remaining}      ★★★"

            result_rows.append({'values': new_values, 'happo': row['happo']})

        result_sheets[sheet_name] = result_rows

    return result_sheets, headers


def save_sheets(result_sheets, headers, order_file_path, save_dir=None):
    base_dir = save_dir if save_dir else os.path.dirname(order_file_path)
    base_name = os.path.splitext(os.path.basename(order_file_path))[0]
    happo_fill = PatternFill(start_color=HAPPO_COLOR, end_color=HAPPO_COLOR, fill_type='solid')

    saved = []
    for sheet_name, rows in result_sheets.items():
        out_path = os.path.join(base_dir, f"{base_name}_{sheet_name}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for i, row in enumerate(rows, start=2):
            ws.append(row['values'])
            ws.row_dimensions[i].height = 16.40
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(wrap_text=True)
            if row['happo']:
                for cell in ws[ws.max_row]:
                    cell.fill = happo_fill

        wb.save(out_path)
        saved.append(out_path)

    return saved


class App(TkinterDnD.Tk if HAS_DND else tk.Tk):
    def __init__(self):
        super().__init__()
        current_ver = get_current_version()
        self.title(f'발주서 처리 프로그램 v{current_ver}')
        self.geometry('720x600')
        self.resizable(False, False)
        self.configure(bg='#f5f5f5')

        self.order_file = tk.StringVar()
        self.location_file = tk.StringVar()
        self.stock_file = tk.StringVar()
        self.save_folder = tk.StringVar()
        self.mapping = load_saved_mapping()
        self._headers = None
        self._rows = None
        self._detected_ids = []
        self.mapping_entries = []

        self.show_step1()

        # 백그라운드에서 업데이트 체크 (1초 후 실행)
        self.after(1000, self._check_update_async)

    def _check_update_async(self):
        """백그라운드 스레드에서 업데이트 체크"""
        def _worker():
            result = check_for_update()
            if result:
                latest, download_url = result
                self.after(0, lambda: self._show_update_dialog(latest, download_url))
        threading.Thread(target=_worker, daemon=True).start()

    def _show_update_dialog(self, latest_version, download_url):
        """업데이트 알림 다이얼로그"""
        current = get_current_version()
        msg = (
            f"새 버전이 출시되었습니다!\n\n"
            f"  · 현재 버전: v{current}\n"
            f"  · 최신 버전: v{latest_version}\n\n"
            f"[다운로드 방법]\n"
            f"  1. '예' 클릭 → 다운로드 페이지 열림\n"
            f"  2. order_processor.exe 다운로드\n"
            f"  3. 경고창 뜨면 '추가 정보' → '실행' 클릭\n"
            f"  4. 또는 파일 우클릭 → 속성 → 차단 해제 체크\n\n"
            f"다운로드 페이지를 여시겠습니까?"
        )
        if messagebox.askyesno('업데이트 알림', msg):
            webbrowser.open(f'https://github.com/{GITHUB_REPO}/releases/latest')

    def _enable_drop(self, widget, var):
        """Entry 위젯에 드래그 앤 드롭 지원 추가"""
        if not HAS_DND:
            return
        def on_drop(event):
            path = event.data.strip()
            # Windows에서 중괄호로 감싸진 경로 처리
            if path.startswith('{') and path.endswith('}'):
                path = path[1:-1]
            var.set(path)
        widget.drop_target_register(DND_FILES)
        widget.dnd_bind('<<Drop>>', on_drop)

    def clear(self):
        for w in self.winfo_children():
            w.destroy()

    def make_header(self, step, title):
        f = tk.Frame(self, bg='#2c3e50', pady=12)
        f.pack(fill='x')
        tk.Label(f, text=f'  STEP {step}  |  {title}',
                 font=('맑은 고딕', 13, 'bold'), bg='#2c3e50', fg='white').pack(anchor='w')
        ind = tk.Frame(self, bg='#ecf0f1', pady=6)
        ind.pack(fill='x')
        steps = ['파일 선택', '매핑 확인', '미리보기', '처리 완료', '출고리스트']
        for i, s in enumerate(steps, 1):
            color = '#27ae60' if i < step else ('#2980b9' if i == step else '#bdc3c7')
            tk.Label(ind, text=f'● {s}', font=('맑은 고딕', 9),
                     bg='#ecf0f1', fg=color).pack(side='left', padx=18)

    # ─── STEP 1: 파일 선택 ────────────────────────────────────────────
    def show_step1(self):
        self.clear()
        self.make_header(1, '파일 선택')

        frame = tk.Frame(self, bg='#f5f5f5', padx=35, pady=25)
        frame.pack(fill='both', expand=True)
        frame.columnconfigure(0, weight=1)

        # ① 위치 파일 (구글시트)
        tk.Label(frame, text='① 위치 파일  (구글시트 자동 다운로드)',
                 font=('맑은 고딕', 11, 'bold'), bg='#f5f5f5').grid(
                 row=0, column=0, sticky='w', pady=(8, 2))
        row1 = tk.Frame(frame, bg='#f5f5f5')
        row1.grid(row=1, column=0, sticky='ew')
        row1.columnconfigure(0, weight=1)
        ent1 = tk.Entry(row1, textvariable=self.location_file, font=('맑은 고딕', 10),
                 relief='solid', bd=1)
        ent1.grid(row=0, column=0, sticky='ew', ipady=4)
        self._enable_drop(ent1, self.location_file)
        tk.Button(row1, text='구글시트 다운로드', command=self._download_gsheet,
                  bg='#27ae60', fg='white', font=('맑은 고딕', 10),
                  padx=10, relief='flat', cursor='hand2').grid(row=0, column=1, padx=(6, 0))
        tk.Button(row1, text='찾아보기', command=lambda: self._browse(self.location_file),
                  bg='#3498db', fg='white', font=('맑은 고딕', 10),
                  padx=12, relief='flat', cursor='hand2').grid(row=0, column=2, padx=(4, 0))
        self.dl_status = tk.Label(frame, text='', font=('맑은 고딕', 9),
                                  bg='#f5f5f5', fg='#27ae60')
        self.dl_status.grid(row=2, column=0, sticky='w', pady=(0, 2))

        # ② 발주서 파일
        tk.Label(frame, text='② 발주서 파일',
                 font=('맑은 고딕', 11, 'bold'), bg='#f5f5f5').grid(
                 row=3, column=0, sticky='w', pady=(10, 2))
        row2 = tk.Frame(frame, bg='#f5f5f5')
        row2.grid(row=4, column=0, sticky='ew')
        row2.columnconfigure(0, weight=1)
        ent2 = tk.Entry(row2, textvariable=self.order_file, font=('맑은 고딕', 10),
                 relief='solid', bd=1)
        ent2.grid(row=0, column=0, sticky='ew', ipady=4)
        self._enable_drop(ent2, self.order_file)
        tk.Button(row2, text='자동복구', command=self._recover_order_file,
                  bg='#e67e22', fg='white', font=('맑은 고딕', 10),
                  padx=10, relief='flat', cursor='hand2').grid(row=0, column=1, padx=(6, 0))
        tk.Button(row2, text='찾아보기', command=lambda: self._browse(self.order_file),
                  bg='#3498db', fg='white', font=('맑은 고딕', 10),
                  padx=12, relief='flat', cursor='hand2').grid(row=0, column=2, padx=(4, 0))

        # ③ 재고 파일 / ④ 저장 폴더
        for row_i, (label, var, mode) in enumerate([
            ('③ 재고 파일  (재고현황조회)', self.stock_file, 'file'),
            ('④ 저장 폴더  (비워두면 발주서 파일과 같은 폴더)', self.save_folder, 'folder'),
        ], start=3):
            tk.Label(frame, text=label, font=('맑은 고딕', 11, 'bold'),
                     bg='#f5f5f5').grid(row=row_i*2, column=0, sticky='w', pady=(10, 2))
            row_f = tk.Frame(frame, bg='#f5f5f5')
            row_f.grid(row=row_i*2+1, column=0, sticky='ew')
            row_f.columnconfigure(0, weight=1)
            ent = tk.Entry(row_f, textvariable=var, font=('맑은 고딕', 10),
                     relief='solid', bd=1)
            ent.grid(row=0, column=0, sticky='ew', ipady=4)
            self._enable_drop(ent, var)
            cmd = (lambda v=var: self._browse_folder(v)) if mode == 'folder' else (lambda v=var: self._browse(v))
            tk.Button(row_f, text='찾아보기', command=cmd,
                      bg='#3498db', fg='white', font=('맑은 고딕', 10),
                      padx=12, relief='flat', cursor='hand2').grid(row=0, column=1, padx=(6, 0))

        btn_frame = tk.Frame(self, bg='#f5f5f5', pady=15)
        btn_frame.pack(fill='x', padx=35)
        tk.Button(btn_frame, text='다음  →', command=self._step1_next,
                  bg='#27ae60', fg='white', font=('맑은 고딕', 12, 'bold'),
                  padx=22, pady=9, relief='flat', cursor='hand2').pack(side='right')

    def _recover_order_file(self):
        path = self.order_file.get().strip()
        if not path:
            messagebox.showwarning('알림', '먼저 발주서 파일을 선택해주세요.')
            return
        if not os.path.exists(path):
            messagebox.showerror('오류', f'파일이 존재하지 않습니다:\n{path}')
            return
        try:
            out_path, method, count = recover_order_file(path)

            # 복구된 파일 데이터 검증
            try:
                recovered_headers, recovered_rows = load_order_file(out_path)
                errors = self._validate_order_data(recovered_headers, recovered_rows)
            except OrderFileError as ve:
                errors = [f"  · 복구된 파일 헤더 검증 실패: {ve}"]

            if errors:
                # 일부 오류 있음 → 부분 복구로 처리
                self.order_file.set(out_path)
                limit = 20
                shown = errors[:limit]
                msg = f"복구는 완료되었으나 {len(errors)}건의 데이터 오류가 남아있습니다.\n"
                msg += f"(아래 {min(limit, len(errors))}건 표시)\n\n"
                msg += '\n'.join(shown)
                if len(errors) > limit:
                    msg += f"\n... 외 {len(errors) - limit}건 더"
                msg += f"\n\n[복구 방법]\n  · {method}\n"
                msg += f"[복구된 행 수]\n  · {count}건\n"
                msg += f"[저장 파일]\n  · {os.path.basename(out_path)}\n\n"
                msg += "오류 부분을 수정한 후 다시 시도해주세요."
                messagebox.showerror('복구 부분 성공 (데이터 오류 남음)', msg)
            else:
                # 100% 깨끗 → 진행 가능
                self.order_file.set(out_path)
                messagebox.showinfo(
                    '복구 완료 (100%)',
                    f'발주서 파일이 완전히 복구되었습니다. 진행 가능합니다.\n\n'
                    f'[복구 방법]\n  · {method}\n\n'
                    f'[복구된 행 수]\n  · {count}건\n\n'
                    f'[저장 파일]\n  · {os.path.basename(out_path)}\n\n'
                    f'※ 발주서 파일 경로가 복구된 파일로 자동 변경되었습니다.\n'
                    f'※ 합포 색상 정보는 유실될 수 있으니 결과를 확인해주세요.'
                )
        except OrderFileError as e:
            messagebox.showerror('복구 실패', str(e))
        except Exception as e:
            import traceback
            messagebox.showerror('복구 오류', f'{e}\n\n{traceback.format_exc()}')

    def _browse(self, var):
        path = filedialog.askopenfilename(filetypes=[('Excel 파일', '*.xlsx *.xls')])
        if path:
            var.set(path)

    def _browse_folder(self, var):
        path = filedialog.askdirectory()
        if path:
            var.set(path)

    def _download_gsheet(self):
        self.dl_status.config(text='Chrome에서 다운로드 중... 잠시 기다려주세요.', fg='#e67e22')
        self.update()

        # 다운로드 전 기존 파일 목록 스냅샷
        before = set(glob.glob(os.path.join(DOWNLOAD_FOLDER, '*.xlsx')))

        # Chrome으로 다운로드 URL 열기 (로그인 세션 사용)
        webbrowser.open(GSHEET_URL)

        # 백그라운드에서 파일 감지
        threading.Thread(target=self._wait_for_download, args=(before,), daemon=True).start()

    def _wait_for_download(self, before):
        timeout = 30
        start = time.time()
        while time.time() - start < timeout:
            time.sleep(1)
            after = set(glob.glob(os.path.join(DOWNLOAD_FOLDER, '*.xlsx')))
            new_files = after - before
            # .crdownload(다운로드 중) 파일 제외
            completed = [f for f in new_files if not f.endswith('.crdownload')]
            if completed:
                newest = max(completed, key=os.path.getmtime)
                self.after(0, lambda f=newest: self._on_download_complete(f))
                return
        self.after(0, lambda: self.dl_status.config(
            text='다운로드 감지 실패. 직접 선택해주세요.', fg='red'))

    def _on_download_complete(self, path):
        from datetime import datetime
        import shutil
        # 바탕화면에 날짜_오전/오후 폴더 생성
        now = datetime.now()
        ampm = '오전' if now.hour < 12 else '오후'
        folder_name = now.strftime('%Y%m%d') + f'_{ampm}'
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        if not os.path.exists(desktop):
            desktop = os.path.join(os.path.expanduser('~'), 'OneDrive', '바탕 화면')
        target_dir = os.path.join(desktop, folder_name)
        os.makedirs(target_dir, exist_ok=True)

        # 파일 이동
        fname = os.path.basename(path)
        new_path = os.path.join(target_dir, fname)
        try:
            shutil.move(path, new_path)
        except Exception:
            new_path = path  # 이동 실패 시 원래 경로 유지

        self.location_file.set(new_path)
        self.save_folder.set(target_dir)
        self.dl_status.config(text=f'✓ 다운로드 완료: {folder_name}/{fname}', fg='#27ae60')

    def _step1_next(self):
        if not self.order_file.get():
            messagebox.showwarning('알림', '발주서 파일을 선택해주세요.')
            return
        if not self.stock_file.get():
            messagebox.showwarning('알림', '재고 파일을 선택해주세요.')
            return
        try:
            self._headers, self._rows = load_order_file(self.order_file.get())

            # 데이터 유효성 검사
            errors = self._validate_order_data(self._headers, self._rows)
            if errors:
                # 오류 많으면 앞 20개만 표시
                limit = 20
                shown = errors[:limit]
                msg = f"발주서 데이터에 {len(errors)}건의 오류가 있습니다.\n"
                msg += f"(아래 {min(limit, len(errors))}건 표시)\n\n"
                msg += '\n'.join(shown)
                if len(errors) > limit:
                    msg += f"\n... 외 {len(errors) - limit}건 더"
                msg += "\n\n발주서 파일에서 해당 셀을 수정한 후 다시 시도해주세요."
                messagebox.showerror('데이터 오류', msg)
                return

            id_col = self._headers.index('아이디')
            ids = sorted(set(
                str(r['values'][id_col]).strip()
                for r in self._rows if r['values'][id_col]
            ))
            self._detected_ids = ids
            for id_ in ids:
                if id_ not in self.mapping:
                    self.mapping[id_] = id_
            self.show_step2()
        except OrderFileError as e:
            messagebox.showerror('발주서 파일 오류', str(e))
        except Exception as e:
            import traceback
            messagebox.showerror('오류', f'파일 읽기 실패:\n{e}\n\n{traceback.format_exc()}')

    def _validate_order_data(self, headers, rows):
        """발주서 데이터 행 단위 검증. 오류 메시지 리스트 반환."""
        errors = []
        # 검사할 필수 컬럼 (필드명 → 설명)
        required_fields = {
            '아이디': '아이디',
            '상품코드': '상품코드',
            '수량': '수량',
            '사방넷 상품명': '사방넷 상품명',
            '수화주명': '수화주명',
            '주소': '주소',
        }
        # 헤더 인덱스 매핑
        col_idx = {}
        for field in required_fields:
            if field in headers:
                col_idx[field] = headers.index(field)

        for i, row in enumerate(rows, start=2):  # 엑셀은 2행부터 데이터 (1행=헤더)
            vals = row['values']

            # 전체 행이 비어있으면 스킵
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in vals):
                continue

            for field, desc in required_fields.items():
                if field not in col_idx:
                    continue
                idx = col_idx[field]
                if idx >= len(vals):
                    errors.append(f"  · {i}행: [{desc}] 열이 없음")
                    continue
                v = vals[idx]
                # 빈 값 체크
                if v is None or (isinstance(v, str) and not v.strip()):
                    errors.append(f"  · {i}행: [{desc}] 값이 비어있음")
                    continue
                # 수량은 숫자여야 함
                if field == '수량':
                    try:
                        qty = int(v)
                        if qty <= 0:
                            errors.append(f"  · {i}행: [{desc}] 0 이하 값 ({v})")
                    except (ValueError, TypeError):
                        errors.append(f"  · {i}행: [{desc}] 숫자가 아님 ({v!r})")

        return errors

    # ─── STEP 2: 매핑 확인 ───────────────────────────────────────────
    def show_step2(self):
        self.clear()
        self.make_header(2, '아이디 ↔ 시트명 매핑 확인')

        frame = tk.Frame(self, bg='#f5f5f5', padx=35, pady=10)
        frame.pack(fill='both', expand=True)

        tk.Label(frame, text='시트명을 확인하고 필요시 수정하세요.  (체크박스로 선택 후 저장/삭제)',
                 font=('맑은 고딕', 10), bg='#f5f5f5', fg='#666').pack(anchor='w', pady=(0, 8))

        hdr = tk.Frame(frame, bg='#34495e')
        hdr.pack(fill='x')
        # 전체선택 체크박스
        self._select_all_var = tk.BooleanVar(value=False)
        tk.Checkbutton(hdr, variable=self._select_all_var, command=self._toggle_select_all,
                       bg='#34495e', activebackground='#34495e').pack(side='left', padx=(8, 0))
        for text, w in [('아이디', 24), ('시트명', 24)]:
            tk.Label(hdr, text=text, font=('맑은 고딕', 10, 'bold'),
                     bg='#34495e', fg='white', width=w, anchor='w',
                     padx=8, pady=5).pack(side='left')

        outer = tk.Frame(frame, bg='#f5f5f5')
        outer.pack(fill='both', expand=True)
        canvas = tk.Canvas(outer, bg='#f5f5f5', highlightthickness=0, height=280)
        scroll = ttk.Scrollbar(outer, orient='vertical', command=canvas.yview)
        inner = tk.Frame(canvas, bg='#f5f5f5')
        inner.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=inner, anchor='nw')
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side='left', fill='both', expand=True)
        scroll.pack(side='right', fill='y')

        self.mapping_entries = []
        self._mapping_inner = inner
        self._mapping_canvas = canvas

        # 저장된 매핑 + 발주서에서 감지된 아이디 모두 표시
        all_ids = list(dict.fromkeys(
            list(self.mapping.keys()) + self._detected_ids
        ))
        for id_ in all_ids:
            self._add_mapping_row(inner, id_, self.mapping.get(id_, id_), editable_id=False)

        # 버튼 프레임
        btn_frame = tk.Frame(self, bg='#f5f5f5', pady=15)
        btn_frame.pack(fill='x', padx=35)
        tk.Button(btn_frame, text='←  이전', command=self.show_step1,
                  font=('맑은 고딕', 11), padx=15, pady=7, relief='flat',
                  bg='#95a5a6', fg='white', cursor='hand2').pack(side='left')
        tk.Button(btn_frame, text='+  추가', command=self._add_manual_mapping,
                  bg='#e67e22', fg='white', font=('맑은 고딕', 11),
                  padx=12, pady=7, relief='flat', cursor='hand2').pack(side='left', padx=(10, 0))
        tk.Button(btn_frame, text='저장', command=self._save_mapping,
                  bg='#2980b9', fg='white', font=('맑은 고딕', 11),
                  padx=12, pady=7, relief='flat', cursor='hand2').pack(side='left', padx=(10, 0))
        tk.Button(btn_frame, text='선택 삭제', command=self._delete_selected_mapping,
                  bg='#c0392b', fg='white', font=('맑은 고딕', 11),
                  padx=12, pady=7, relief='flat', cursor='hand2').pack(side='left', padx=(10, 0))
        tk.Button(btn_frame, text='다음  →', command=self._step2_next,
                  bg='#27ae60', fg='white', font=('맑은 고딕', 12, 'bold'),
                  padx=22, pady=9, relief='flat', cursor='hand2').pack(side='right')

    def _add_mapping_row(self, inner, id_text, sheet_text, editable_id=False):
        i = len(self.mapping_entries)
        bg = '#ffffff' if i % 2 == 0 else '#f8f9fa'
        row_f = tk.Frame(inner, bg=bg)
        row_f.pack(fill='x')
        # 체크박스
        chk_var = tk.BooleanVar(value=False)
        tk.Checkbutton(row_f, variable=chk_var, bg=bg,
                       activebackground=bg).pack(side='left', padx=(8, 0))
        if editable_id:
            id_var = tk.StringVar(value=id_text)
            tk.Entry(row_f, textvariable=id_var, font=('맑은 고딕', 10),
                     width=24, relief='solid', bd=1).pack(side='left', padx=(4, 0), pady=3)
        else:
            id_var = None
            tk.Label(row_f, text=id_text, font=('맑은 고딕', 10), bg=bg,
                     width=24, anchor='w', padx=8, pady=5).pack(side='left')
        sheet_var = tk.StringVar(value=sheet_text)
        tk.Entry(row_f, textvariable=sheet_var, font=('맑은 고딕', 10),
                 width=24, relief='solid', bd=1).pack(side='left', padx=6, pady=3)
        self.mapping_entries.append((id_text, sheet_var, id_var, chk_var, row_f))

    def _toggle_select_all(self):
        val = self._select_all_var.get()
        for entry in self.mapping_entries:
            entry[3].set(val)

    def _add_manual_mapping(self):
        self._add_mapping_row(self._mapping_inner, '', '', editable_id=True)
        self._mapping_canvas.update_idletasks()
        self._mapping_canvas.yview_moveto(1.0)

    def _save_mapping(self):
        """현재 매핑 목록을 파일로 영구 저장"""
        new_mapping = {}
        for entry in self.mapping_entries:
            id_text, sheet_var, id_var, chk_var, row_f = entry
            if id_var is not None:
                id_text = id_var.get().strip()
            if not id_text:
                continue
            name = sheet_var.get().strip()
            new_mapping[id_text] = name if name else id_text
        self.mapping = new_mapping
        save_mapping_file(self.mapping)
        messagebox.showinfo('저장 완료', f'매핑 {len(new_mapping)}건이 저장되었습니다.')

    def _delete_selected_mapping(self):
        """선택된 매핑 행 삭제"""
        selected = [e for e in self.mapping_entries if e[3].get()]
        if not selected:
            messagebox.showwarning('알림', '삭제할 항목을 선택해주세요.')
            return
        if not messagebox.askyesno('삭제 확인', f'{len(selected)}건을 삭제하시겠습니까?'):
            return
        for entry in selected:
            id_text, sheet_var, id_var, chk_var, row_f = entry
            if id_var is not None:
                id_text = id_var.get().strip()
            if id_text in self.mapping:
                del self.mapping[id_text]
            row_f.destroy()
        self.mapping_entries = [e for e in self.mapping_entries if not e[3].get()]
        save_mapping_file(self.mapping)

    def _step2_next(self):
        for entry in self.mapping_entries:
            id_text, sheet_var, id_var = entry[0], entry[1], entry[2]
            if id_var is not None:
                id_text = id_var.get().strip()
            if not id_text:
                continue
            name = sheet_var.get().strip()
            self.mapping[id_text] = name if name else id_text
        self.show_step3()

    # ─── STEP 3: 미리보기 ────────────────────────────────────────────
    def show_step3(self):
        self.clear()
        self.make_header(3, '미리보기 확인')

        frame = tk.Frame(self, bg='#f5f5f5', padx=35, pady=15)
        frame.pack(fill='both', expand=True)

        try:
            id_col = self._headers.index('아이디')
            stats = defaultdict(lambda: {'rows': 0, 'happo': 0, 'overseas': 0, 'damaged': 0})
            for row in self._rows:
                id_ = str(row['values'][id_col]).strip() if row['values'][id_col] else '기타'
                sheet = self.mapping.get(id_, id_)
                stats[sheet]['rows'] += 1
                if row['happo']:
                    stats[sheet]['happo'] += 1
                if row['overseas']:
                    stats[sheet]['overseas'] += 1
                if row['damaged']:
                    stats[sheet]['damaged'] += 1

            tk.Label(frame, text='처리 후 생성될 파일 목록입니다.',
                     font=('맑은 고딕', 10), bg='#f5f5f5', fg='#666').pack(anchor='w', pady=(0, 10))

            # Treeview 스타일
            style = ttk.Style()
            style.theme_use('clam')
            style.configure('Preview.Treeview',
                            font=('맑은 고딕', 10),
                            rowheight=28,
                            background='#ffffff',
                            fieldbackground='#ffffff',
                            borderwidth=0)
            style.configure('Preview.Treeview.Heading',
                            font=('맑은 고딕', 10, 'bold'),
                            background='#34495e',
                            foreground='white',
                            borderwidth=0,
                            relief='flat')
            style.map('Preview.Treeview.Heading',
                       background=[('active', '#2c3e50')])
            style.map('Preview.Treeview',
                       background=[('selected', '#d5e8f0')])

            columns = ('file', 'orders', 'happo', 'overseas', 'damaged')
            tree = ttk.Treeview(frame, columns=columns, show='headings',
                                style='Preview.Treeview',
                                height=min(len(stats) + 1, 10))

            tree.heading('file', text='저장 파일명', anchor='w')
            tree.heading('orders', text='주문', anchor='center')
            tree.heading('happo', text='합포', anchor='center')
            tree.heading('overseas', text='해외', anchor='center')
            tree.heading('damaged', text='훼손', anchor='center')

            tree.column('file', width=340, minwidth=200, anchor='w')
            tree.column('orders', width=70, minwidth=60, anchor='center')
            tree.column('happo', width=70, minwidth=60, anchor='center')
            tree.column('overseas', width=70, minwidth=60, anchor='center')
            tree.column('damaged', width=70, minwidth=60, anchor='center')

            # 줄무늬 태그
            tree.tag_configure('even', background='#ffffff')
            tree.tag_configure('odd', background='#f0f7ff')
            tree.tag_configure('total', background='#34495e', foreground='white',
                               font=('맑은 고딕', 10, 'bold'))

            base = os.path.splitext(os.path.basename(self.order_file.get()))[0]
            total_rows = 0
            total_happo = 0
            total_overseas = 0
            total_damaged = 0
            for i, (sheet, s) in enumerate(stats.items()):
                fname = f"{base}_{sheet}.xlsx"
                tag = 'even' if i % 2 == 0 else 'odd'
                tree.insert('', 'end', values=(
                    fname, f"{s['rows']}건", f"{s['happo']}건",
                    f"{s['overseas']}건", f"{s['damaged']}건"
                ), tags=(tag,))
                total_rows += s['rows']
                total_happo += s['happo']
                total_overseas += s['overseas']
                total_damaged += s['damaged']

            # 합계 행
            tree.insert('', 'end', values=(
                '합계', f"{total_rows}건", f"{total_happo}건",
                f"{total_overseas}건", f"{total_damaged}건"
            ), tags=('total',))

            tree.pack(fill='x', pady=(0, 10))

            tk.Label(frame, text=f'총 {total_rows}건  \u2192  {len(stats)}개 파일 생성',
                     font=('맑은 고딕', 11, 'bold'), bg='#f5f5f5', fg='#27ae60').pack(anchor='w')

        except Exception as e:
            tk.Label(frame, text=f'오류: {e}', bg='#f5f5f5', fg='red').pack()

        btn_frame = tk.Frame(self, bg='#f5f5f5', pady=15)
        btn_frame.pack(fill='x', padx=35)
        tk.Button(btn_frame, text='←  이전', command=self.show_step2,
                  font=('맑은 고딕', 11), padx=15, pady=7, relief='flat',
                  bg='#95a5a6', fg='white', cursor='hand2').pack(side='left')
        tk.Button(btn_frame, text='실행  ▶', command=self.show_step4,
                  bg='#e74c3c', fg='white', font=('맑은 고딕', 12, 'bold'),
                  padx=22, pady=9, relief='flat', cursor='hand2').pack(side='right')

    # ─── STEP 4: 처리 실행 ───────────────────────────────────────────
    def show_step4(self):
        self.clear()
        self.make_header(4, '처리 중')

        frame = tk.Frame(self, bg='#f5f5f5', padx=35, pady=30)
        frame.pack(fill='both', expand=True)

        self.status_label = tk.Label(frame, text='재고 파일 읽는 중...',
                                     font=('맑은 고딕', 12), bg='#f5f5f5')
        self.status_label.pack(pady=(10, 5))

        self.progress = ttk.Progressbar(frame, length=580, mode='determinate')
        self.progress.pack(pady=10)

        self.result_label = tk.Label(frame, text='', font=('맑은 고딕', 10),
                                     bg='#f5f5f5', fg='#27ae60',
                                     wraplength=620, justify='left')
        self.result_label.pack(pady=15)

        self.after(150, self._run_process)

    def _run_process(self):
        try:
            self.status_label.config(text='위치 파일 읽는 중...')
            self.progress['value'] = 10
            self.update()

            location_map = {}
            if self.location_file.get():
                try:
                    location_map = load_location_file(self.location_file.get())
                except Exception as e:
                    messagebox.showwarning('위치 파일 오류', f'위치 파일을 읽을 수 없습니다:\n{e}')

            self.status_label.config(text='재고 파일 읽는 중...')
            self.progress['value'] = 25
            self.update()

            stock = load_stock_file(self.stock_file.get())

            self.status_label.config(text='데이터 처리 중...')
            self.progress['value'] = 45
            self.update()

            result_sheets, headers = process_data(
                self._headers, self._rows, self.mapping, stock, location_map)

            self.status_label.config(text='파일 저장 중...')
            self.progress['value'] = 75
            self.update()

            save_dir = self.save_folder.get().strip() or os.path.dirname(self.order_file.get())
            saved = save_sheets(result_sheets, headers, self.order_file.get(), save_dir)

            self.progress['value'] = 100
            self.make_header(4, '처리 완료  ✓')
            self.status_label.config(text='모든 파일이 저장되었습니다.', fg='#27ae60')

            lines = '\n'.join(f'  ✓  {os.path.basename(f)}' for f in saved)
            self.result_label.config(text=f'저장된 파일 ({len(saved)}개):\n{lines}')

            # 출고리스트용 데이터 저장
            self._result_sheets = result_sheets
            self._stock = stock
            self._location_map = location_map

            folder = self.save_folder.get().strip() or os.path.dirname(self.order_file.get())
            btn_frame = tk.Frame(self, bg='#f5f5f5', pady=15)
            btn_frame.pack(fill='x', padx=35)
            tk.Button(btn_frame, text='저장 폴더 열기', command=lambda: os.startfile(folder),
                      bg='#3498db', fg='white', font=('맑은 고딕', 11),
                      padx=15, pady=7, relief='flat', cursor='hand2').pack(side='left')
            tk.Button(btn_frame, text='출고리스트 생성  →', command=self.show_step5,
                      bg='#8e44ad', fg='white', font=('맑은 고딕', 12, 'bold'),
                      padx=22, pady=9, relief='flat', cursor='hand2').pack(side='right')

        except Exception as e:
            self.status_label.config(text='오류 발생', fg='red')
            messagebox.showerror('처리 오류', str(e))

    # ─── STEP 5: 출고리스트 생성 ─────────────────────────────────────
    def show_step5(self):
        self.clear()
        self.make_header(5, '출고리스트 생성 중')

        frame = tk.Frame(self, bg='#f5f5f5', padx=35, pady=30)
        frame.pack(fill='both', expand=True)

        self.status_label = tk.Label(frame, text='출고리스트 생성 중...',
                                     font=('맑은 고딕', 12), bg='#f5f5f5')
        self.status_label.pack(pady=(10, 5))

        self.progress = ttk.Progressbar(frame, length=580, mode='determinate')
        self.progress.pack(pady=10)

        self.result_label = tk.Label(frame, text='', font=('맑은 고딕', 10),
                                     bg='#f5f5f5', fg='#27ae60',
                                     wraplength=620, justify='left')
        self.result_label.pack(pady=15)

        self.after(150, self._run_shiplist)

    def _run_shiplist(self):
        try:
            self.progress['value'] = 20
            self.status_label.config(text='출고리스트 데이터 집계 중...')
            self.update()

            headers = self._headers
            code_col = headers.index('상품코드')
            name_col = headers.index('사방넷 상품명')
            qty_col = headers.index('수량')

            sheet_names = list(self._result_sheets.keys())

            # 거래처별 상품코드 집계: {sheet: {code: {name, location, normal, happo}}}
            sheet_data = {}
            for sheet_name, rows in self._result_sheets.items():
                code_map = {}
                for row in rows:
                    code = str(row['values'][code_col]).strip() if row['values'][code_col] else None
                    if not code:
                        continue
                    raw_name = str(row['values'][name_col] or '')
                    # ★★★ 줄 제거
                    product_name = raw_name.split('\n')[0].strip()
                    qty = int(row['values'][qty_col] or 0)

                    if code not in code_map:
                        # 상품위치: location_map에 있으면 해당 위치, 없으면 '0'
                        location = self._location_map.get(code, '0') if self._location_map else '0'
                        code_map[code] = {
                            'name': product_name,
                            'location': location,
                            'normal': 0,
                            'happo': 0,
                        }
                    if row['happo']:
                        code_map[code]['happo'] += qty
                    else:
                        code_map[code]['normal'] += qty

                sheet_data[sheet_name] = code_map

            self.progress['value'] = 50
            self.status_label.config(text='엑셀 파일 생성 중...')
            self.update()

            # 출고리스트 엑셀 생성
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 기본 시트 제거

            from datetime import datetime
            today = datetime.now().strftime('%Y-%m-%d')
            ampm = '오전' if datetime.now().hour < 12 else '오후'

            header_fill = PatternFill(start_color='FF34495E', end_color='FF34495E', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            title_font = Font(bold=True, size=14)
            data_font = Font(size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )

            stock = self._stock

            for idx, sheet_name in enumerate(sheet_names):
                ws = wb.create_sheet(title=f'{sheet_name}출고리스트')
                code_map = sheet_data.get(sheet_name, {})
                other_sheets = [s for s in sheet_names if s != sheet_name]

                # 행1: 날짜 + 오전/오후 + 거래처명
                ws.append([today, f'{ampm} {sheet_name}'])
                ws['A1'].font = title_font
                ws['B1'].font = title_font

                # 행2: 헤더
                col_headers = ['상품코드', '상품명', '상품위치', '출고', '합포장', '재고'] + other_sheets
                ws.append(col_headers)
                for cell in ws[2]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                # 행3~: 데이터 (상품코드 기준 정렬)
                for code in sorted(code_map.keys(), key=lambda c: code_map[c]['location']):
                    info = code_map[code]
                    ship_qty = info['normal'] + info['happo']
                    happo_qty = info['happo']
                    avail = stock.get(code, 0)
                    # 전체 거래처 출고수량 합산하여 잔여재고 계산
                    total_ship = sum(
                        sd.get(code, {}).get('normal', 0) + sd.get(code, {}).get('happo', 0)
                        for sd in sheet_data.values()
                    )
                    remaining = avail - total_ship

                    row_data = [
                        code,
                        info['name'],
                        info['location'],
                        ship_qty if ship_qty > 0 else '',
                        happo_qty if happo_qty > 0 else '',
                        remaining,
                    ]

                    # 다른 거래처 출고수량
                    for other in other_sheets:
                        other_info = sheet_data.get(other, {}).get(code)
                        if other_info:
                            other_qty = other_info['normal'] + other_info['happo']
                            row_data.append(other_qty if other_qty > 0 else '')
                        else:
                            row_data.append('')

                    ws.append(row_data)
                    for cell in ws[ws.max_row]:
                        cell.border = thin_border
                        cell.font = data_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    # 상품명은 왼쪽 정렬
                    ws.cell(ws.max_row, 2).alignment = Alignment(horizontal='left', vertical='center')
                    # 행 높이
                    ws.row_dimensions[ws.max_row].height = 24

                # 행 높이 (헤더 포함)
                ws.row_dimensions[1].height = 28
                ws.row_dimensions[2].height = 24

                # 열 너비 조정
                ws.column_dimensions['A'].width = 14
                ws.column_dimensions['B'].width = 40
                ws.column_dimensions['C'].width = 9
                ws.column_dimensions['D'].width = 7
                ws.column_dimensions['E'].width = 7
                ws.column_dimensions['F'].width = 7
                for i, other_name in enumerate(other_sheets):
                    col_letter = openpyxl.utils.get_column_letter(7 + i)
                    # 회사명 길이 기준으로 열 너비 자동 조정 (최소 7, 최대 20)
                    auto_width = max(7, min(20, len(other_name) * 2 + 2))
                    ws.column_dimensions[col_letter].width = auto_width

                # 인쇄 설정: A4 가로, 한 장에 약 40행
                ws.page_setup.paperSize = ws.PAPERSIZE_A4  # A4
                ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 세로
                ws.page_setup.fitToWidth = 1  # 너비 1페이지
                ws.page_setup.fitToHeight = 0  # 높이는 자동 (여러 페이지)
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                # 여백 최소화 (단위: 인치)
                ws.page_margins.left = 0.3
                ws.page_margins.right = 0.3
                ws.page_margins.top = 0.3
                ws.page_margins.bottom = 0.3
                ws.page_margins.header = 0.15
                ws.page_margins.footer = 0.15
                # 가운데 정렬 (가로)
                ws.print_options.horizontalCentered = True
                # 헤더(행1~2) 매 페이지 반복, 페이지 번호 표시
                ws.print_title_rows = '1:2'
                ws.oddFooter.center.text = '&P / &N'

                self.progress['value'] = 50 + int(40 * (idx + 1) / len(sheet_names))
                self.update()

            # 저장
            save_dir = self.save_folder.get().strip() or os.path.dirname(self.order_file.get())
            base_name = os.path.splitext(os.path.basename(self.order_file.get()))[0]
            out_path = os.path.join(save_dir, f'{base_name}_출고리스트.xlsx')
            wb.save(out_path)

            self.progress['value'] = 100
            self.make_header(5, '출고리스트 생성 완료  ✓')
            self.status_label.config(text='출고리스트가 생성되었습니다.', fg='#27ae60')

            sheet_list = '\n'.join(f'  ✓  {s}출고리스트' for s in sheet_names)
            self.result_label.config(
                text=f'저장 파일: {os.path.basename(out_path)}\n\n시트 ({len(sheet_names)}개):\n{sheet_list}')

            folder = save_dir
            btn_frame = tk.Frame(self, bg='#f5f5f5', pady=15)
            btn_frame.pack(fill='x', padx=35)
            tk.Button(btn_frame, text='저장 폴더 열기', command=lambda: os.startfile(folder),
                      bg='#3498db', fg='white', font=('맑은 고딕', 11),
                      padx=15, pady=7, relief='flat', cursor='hand2').pack(side='left')
            tk.Button(btn_frame, text='처음으로', command=self.show_step1,
                      bg='#7f8c8d', fg='white', font=('맑은 고딕', 11),
                      padx=15, pady=7, relief='flat', cursor='hand2').pack(side='right')

        except Exception as e:
            import traceback
            self.status_label.config(text='오류 발생', fg='red')
            messagebox.showerror('출고리스트 오류', f'{e}\n\n{traceback.format_exc()}')


if __name__ == '__main__':
    app = App()
    app.mainloop()
